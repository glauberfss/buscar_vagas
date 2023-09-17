# Primeira vez fazendo isso, espero que se tiver algo de errado vcs possam me ajudar com algo que falta ou até mesmo melhorar!

from selenium import webdriver # fazer automações
from selenium.webdriver.common.keys import Keys # simular o uso do teclado
from selenium.webdriver.common.by import By # permite encontrar elementos
from time import sleep
import openpyxl

nome_vaga = 'Estágio ti'

# entrar no site da https://br.indeed.com/
driver = webdriver.Chrome()
driver.get('https://br.indeed.com/')
sleep(20)

# digitar o nome 
campo_vaga = driver.find_element(By.XPATH, "//input[@id='text-input-what']")
campo_vaga.send_keys('Estágio ti')

# clicar em pesquisar
botao_pesquisar = driver.find_element(By.XPATH, "//div//button[@class='yosegi-InlineWhatWhere-primaryButton']")
botao_pesquisar.click()
sleep(5)

# extrair todas as informações
vagas_encontradas = driver.find_elements(By.XPATH, "//div[@id='mosaic-jobResults']//td[@class='resultContent']")
lista_vagas = []
for encontrada in vagas_encontradas:
    lista_vagas.append(encontrada.text)

try:
    # Extrair todas as informações
    vagas_encontradas = driver.find_elements(By.XPATH, "//div[@id='mosaic-jobResults']//td[@class='resultContent']")
    lista_vagas = [encontrada.text for encontrada in vagas_encontradas]

    # Guardar tudo no Excel
    try:
        # Tentar carregar o workbook existente
        workbook = openpyxl.load_workbook('dados.xlsx')
    except FileNotFoundError:
        # Se o arquivo não existe, criar um novo
        workbook = openpyxl.Workbook()

    # Verificar se a planilha 'Vagas Encontradas' já existe
    if 'Vagas Encontradas' not in workbook.sheetnames:
        # Se não existe, criar uma nova
        workbook.create_sheet('Vagas Encontradas')

    # Acessar a página
    pagina = workbook['Vagas Encontradas']

    # Adicionar cabeçalho
    pagina['A1'] = 'Lista de Vagas'

# Adicionar vagas
    for index, vaga in enumerate(lista_vagas, start=2):
        pagina.cell(row=index, column=1, value=vaga)

    # Salvar o arquivo
    workbook.save('dados.xlsx')

except Exception as error:
    print(f'Ocorreu um erro: {error}')

# Fechar o navegador
driver.close()
